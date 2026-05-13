#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Standardize an input Excel to the project's CSV template header order.

Usage:
  python scripts/standardize_excel_to_template.py --in "/path/to/your/input_data.xlsx" \
       --template app/static/templates/is_element_template_full.csv \
       --out app/static/templates/is_element_template_standardized.xlsx

The script will try to map source headers to template headers using exact and normalized matches
and a small alias dictionary. Missing template columns will be filled with empty strings.
It writes a mapping report: scripts/standardize_report.json and prints a short preview.
"""
import argparse
import csv
import json
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook


def normalize(s):
    if s is None:
        return ''
    s = str(s).strip().lower()
    # replace a bunch of punctuation/spaces with underscore
    s = re.sub(r"[^0-9a-z]+", '_', s)
    s = re.sub(r'_+', '_', s)
    return s.strip('_')

# common aliases mapping (from observed legacy headers)
ALIASES = {
    'synonyms': 'synomyns',
    'synonym_s': 'synomyns',
    'lengthaa': 'length_aa',
    'length_aa': 'length_aa',
    'length.1': 'length_aa',
    'orf1_sequence': 'orf_1_sequence',
    'orf2_sequence': 'orf_2_sequence',
    'related_elements': 'related_element_s',
    'group_name': 'group',
}


def load_template_headers(template_csv_path):
    with open(template_csv_path, 'r', encoding='utf-8') as f:
        r = csv.reader(f)
        row = next(r)
        return [c.strip() for c in row]


def build_mapping(src_headers, tpl_headers):
    # returns mapping dict: tpl_header -> src_header_or_None
    src_norm = {normalize(h): h for h in src_headers}
    mapping = {}
    for th in tpl_headers:
        tnorm = normalize(th)
        chosen = None
        # exact normalized match
        if tnorm in src_norm:
            chosen = src_norm[tnorm]
        else:
            # alias check
            if tnorm in ALIASES:
                alias = ALIASES[tnorm]
                if normalize(alias) in src_norm:
                    chosen = src_norm[normalize(alias)]
            # try reversing: find src whose normalized equals alias value
            if chosen is None:
                # try direct alias matching in source
                for a, b in ALIASES.items():
                    if b == th or b == tnorm:
                        if a in src_norm:
                            chosen = src_norm[a]
                            break
            # fuzzy: find source header that startswith or contains words
            if chosen is None:
                for snorm, sh in src_norm.items():
                    if snorm == tnorm:
                        chosen = sh
                        break
                if chosen is None:
                    for snorm, sh in src_norm.items():
                        if tnorm in snorm or snorm in tnorm:
                            chosen = sh
                            break
        mapping[th] = chosen
    return mapping


def standardize(in_xlsx, template_csv, out_xlsx):
    in_xlsx = Path(in_xlsx)
    template_csv = Path(template_csv)
    out_xlsx = Path(out_xlsx)

    if not in_xlsx.exists():
        raise SystemExit(f"Input file not found: {in_xlsx}")
    if not template_csv.exists():
        raise SystemExit(f"Template CSV not found: {template_csv}")

    tpl_headers = load_template_headers(template_csv)

    wb = load_workbook(in_xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("Input XLSX has no rows")
    src_headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    mapping = build_mapping(src_headers, tpl_headers)

    # create new workbook and write headers (template order)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.append(tpl_headers)

    report = {
        'input_file': str(in_xlsx),
        'template': str(template_csv),
        'output_file': str(out_xlsx),
        'template_headers_count': len(tpl_headers),
        'source_headers_count': len(src_headers),
        'mapping': mapping,
        'rows_in': max(0, len(rows)-1),
        'rows_out': 0,
        'notes': []
    }

    # process data rows
    for rnum, vals in enumerate(rows[1:], start=2):
        vals = list(vals)
        row_dict = {src_headers[i]: (vals[i] if i < len(vals) else None) for i in range(len(src_headers))}
        out_row = []
        for th in tpl_headers:
            sh = mapping.get(th)
            if sh is None:
                out_row.append('')
            else:
                v = row_dict.get(sh, '')
                # normalize None to empty string
                out_row.append('' if v is None else v)
        out_ws.append(out_row)
        report['rows_out'] += 1

    out_wb.save(out_xlsx)
    # write report
    rep_path = Path('scripts/standardize_report.json')
    rep_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding='utf-8')
    return report


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--in', dest='infile', required=True)
    parser.add_argument('--template', dest='template', default='app/static/templates/is_element_template_full.csv')
    parser.add_argument('--out', dest='out', default='app/static/templates/is_element_template_standardized.xlsx')
    args = parser.parse_args()
    r = standardize(args.infile, args.template, args.out)
    print('WROTE:', r['output_file'])
    print('ROWS_IN:', r['rows_in'], 'ROWS_OUT:', r['rows_out'])
    print('MAPPING sample (first 10):')
    for k in list(r['mapping'].keys())[:10]:
        print(k, '=>', r['mapping'][k])
    print('\nReport written to scripts/standardize_report.json')
