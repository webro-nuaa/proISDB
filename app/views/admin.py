# -*- coding: utf-8 -*-
"""
Admin views.
"""
from datetime import datetime, timezone, timedelta
from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, jsonify
from flask_login import login_required, current_user, login_user

from app import db
from app.models import User, ISElement, KnowledgeArticle, PageView, AdminLog, SubmissionHistory, DownloadRequest, DownloadRequestItem
from app.models.system import SearchLog
from app.forms.auth import LoginForm
from app.forms.submission import ISElementSubmissionForm
from app.forms.admin import DeleteForm
from app.utils.helpers import escape_like

admin = Blueprint('admin', __name__)

@admin.route('/login', methods=['GET', 'POST'])
def login():
    """Admin standalone login page."""
    if current_user.is_authenticated and current_user.has_admin_permission():
        return redirect(url_for('admin.index'))
    
    form = LoginForm()
    
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.verify_password(form.password.data):
            if user.has_admin_permission():
                login_user(user, remember=form.remember_me.data)
                
                # Record admin login
                AdminLog.log_action(
                    user=user,
                    action='admin_login',
                    resource_type='system',
                    details={'login_ip': request.remote_addr},
                    request=request
                )
                
                next_page = request.args.get('next')
                if next_page and next_page.startswith('/admin'):
                    return redirect(next_page)
                return redirect(url_for('admin.index'))
            else:
                flash('您不是管理员用户。', 'error')
        else:
            flash('用户名或密码错误。', 'error')
    
    return render_template('admin/login.html', form=form)

@admin.route('/')
@login_required
def index():
    """Admin dashboard."""
    if not current_user.has_admin_permission():
        flash('您没有权限访问此页面。', 'error')
        return redirect(url_for('admin.login'))

    # Gather statistics
    stats = {
        'total_elements': ISElement.query.count(),
        'pending_elements': ISElement.query.filter_by(status='pending').count(),
        'approved_elements': ISElement.query.filter_by(status='approved').count(),
        'rejected_elements': ISElement.query.filter_by(status='rejected').count(),
        'total_articles': KnowledgeArticle.query.count(),
        'published_articles': KnowledgeArticle.query.filter_by(status='published').count(),
        'total_users': User.query.count(),
        'admin_users': User.query.filter_by(role='admin').count()
    }

    # Recent pending submissions
    recent_submissions = ISElement.query.filter_by(status='pending')\
                                       .order_by(ISElement.submission_date.desc())\
                                       .limit(5).all()

    # Recent articles
    recent_articles = KnowledgeArticle.query\
                                     .order_by(KnowledgeArticle.created_at.desc())\
                                     .limit(5).all()

    # Today's page views
    today = datetime.now(timezone.utc).date()
    today_views = PageView.query.filter(
        db.func.date(PageView.created_at) == today
    ).count()
    
    # Access trends for the past 7 days
    from sqlalchemy import func

    week_views = db.session.query(
        func.date(PageView.created_at).label('date'),
        func.count(PageView.id).label('count')
    ).filter(
        PageView.created_at >= datetime.now(timezone.utc) - timedelta(days=7)
    ).group_by(
        func.date(PageView.created_at)
    ).order_by('date').all()
    
    return render_template('admin/index.html',
                         stats=stats,
                         recent_submissions=recent_submissions,
                         recent_articles=recent_articles,
                         today_views=today_views,
                         week_views=week_views)

@admin.route('/is-elements')
@login_required
def is_elements():
    """IS Element data management."""
    if not current_user.has_admin_permission():
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('admin.login'))

    is_advanced = request.args.get('advanced') == '1'

    search = request.args.get('search', '').strip()
    status = request.args.get('status', '')
    family = request.args.get('family', '')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)

    if per_page not in [10, 20, 50, 100]:
        per_page = 20

    query = ISElement.query
    
    if is_advanced:
        # Advanced search
        search_params = {}
        
        # IS Name search
        name = request.args.get('name', '').strip()
        name_match = request.args.get('name_match', 'contains')
        if name:
            safe_name = escape_like(name)
            search_params['name'] = name
            search_params['name_match'] = name_match
            if name_match == 'contains':
                query = query.filter(ISElement.name.contains(safe_name))
            elif name_match == 'begin_with':
                query = query.filter(ISElement.name.startswith(safe_name))
            elif name_match == 'end_with':
                query = query.filter(ISElement.name.endswith(safe_name))
            elif name_match == 'equal_to':
                query = query.filter(ISElement.name == name)
        
        # IS Family search
        family_field = request.args.get('family_field', '').strip()
        family_match = request.args.get('family_match', 'contains')
        if family_field:
            safe_family = escape_like(family_field)
            search_params['family_field'] = family_field
            search_params['family_match'] = family_match
            if family_match == 'contains':
                query = query.filter(ISElement.family.contains(safe_family))
            elif family_match == 'begin_with':
                query = query.filter(ISElement.family.startswith(safe_family))
            elif family_match == 'end_with':
                query = query.filter(ISElement.family.endswith(safe_family))
            elif family_match == 'equal_to':
                query = query.filter(ISElement.family == family_field)
        
        # Origin search
        origin = request.args.get('origin', '').strip()
        origin_match = request.args.get('origin_match', 'contains')
        if origin:
            safe_origin = escape_like(origin)
            search_params['origin'] = origin
            search_params['origin_match'] = origin_match
            if origin_match == 'contains':
                query = query.filter(ISElement.origin.contains(safe_origin))
            elif origin_match == 'begin_with':
                query = query.filter(ISElement.origin.startswith(safe_origin))
            elif origin_match == 'end_with':
                query = query.filter(ISElement.origin.endswith(safe_origin))
            elif origin_match == 'equal_to':
                query = query.filter(ISElement.origin == origin)
        
        # MGE Type search
        mge_type = request.args.get('mge_type', '').strip()
        mge_type_match = request.args.get('mge_type_match', 'contains')
        if mge_type:
            safe_mge = escape_like(mge_type)
            search_params['mge_type'] = mge_type
            search_params['mge_type_match'] = mge_type_match
            if mge_type_match == 'contains':
                query = query.filter(ISElement.mge_type.contains(safe_mge))
            elif mge_type_match == 'begin_with':
                query = query.filter(ISElement.mge_type.startswith(safe_mge))
            elif mge_type_match == 'end_with':
                query = query.filter(ISElement.mge_type.endswith(safe_mge))
            elif mge_type_match == 'equal_to':
                query = query.filter(ISElement.mge_type == mge_type)
        
        # Status filter
        if status:
            search_params['status'] = status
            query = query.filter(ISElement.status == status)
        
        # Host search
        host = request.args.get('host', '').strip()
        if host:
            search_params['host'] = host
            query = query.filter(ISElement.host.contains(escape_like(host)))

        accession = request.args.get('accession', '').strip()
        if accession:
            search_params['accession'] = accession
            query = query.filter(ISElement.accession_number.contains(escape_like(accession)))
        
        # Length search (protein length)
        length_field = request.args.get('length_field', '').strip()
        length_match = request.args.get('length_match', 'equal_to')
        if length_field:
            try:
                length_value = int(length_field)
                search_params['length_field'] = length_field
                search_params['length_match'] = length_match
                if length_match == 'equal_to':
                    query = query.filter(ISElement.length == length_value)
                elif length_match == 'gte':
                    query = query.filter(ISElement.length >= length_value)
                elif length_match == 'lte':
                    query = query.filter(ISElement.length <= length_value)
            except ValueError:
                pass
        
        
        search_params['per_page'] = per_page
        
    else:
        # Simple search
        search_params = {}
        if search:
            safe_search = escape_like(search)
            use_fulltext = len(search) >= 3
            if use_fulltext:
                try:
                    import re
                    sanitized = re.sub(r'[<>"\'\\()@~*+()-]', ' ', search)
                    sanitized = ' '.join(sanitized.split())
                    ft_query = query.filter(
                        db.text("MATCH(name, family, `group`, host, comment) AGAINST(:kw IN BOOLEAN MODE)")
                    ).params(kw=sanitized)
                    if ft_query.count() == 0:
                        use_fulltext = False
                    else:
                        query = ft_query
                except Exception:
                    use_fulltext = False

            if not use_fulltext:
                query = query.filter(
                    db.or_(
                        ISElement.name.contains(safe_search),
                        ISElement.family.contains(safe_search),
                        ISElement.host.contains(safe_search),
                        ISElement.group.contains(safe_search)
                    )
                )
        
        if status:
            query = query.filter(ISElement.status == status)
            
        if family:
            query = query.filter(ISElement.family == family)
    
    # Paginate
    elements = query.order_by(ISElement.id.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    
    # Available families
    available_families = db.session.query(ISElement.family).filter(
        ISElement.family.isnot(None), 
        ISElement.family != ''
    ).distinct().all()
    available_families = [f[0] for f in available_families]
    
    # Stats
    stats = {
        'approved': ISElement.query.filter_by(status='approved').count(),
        'pending': ISElement.query.filter_by(status='pending').count(),
        'rejected': ISElement.query.filter_by(status='rejected').count()
    }
    
    # Delete form instance
    delete_form = DeleteForm()
    
    return render_template('admin/is_elements.html', 
                         elements=elements,
                         current_search=search,
                         current_status=status,
                         current_family=family,
                         per_page=per_page,
                         available_families=available_families,
                         stats=stats,
                         delete_form=delete_form,
                         is_advanced=is_advanced,
                         search_params=search_params)

@admin.route('/is-elements/import', methods=['GET', 'POST'])
@login_required
def import_is_elements():
    """Batch import IS elements from CSV/XLSX"""
    if not current_user.has_admin_permission():
        flash('You do not have permission to access this page.', 'error')
        return redirect(url_for('admin.login'))

    if request.method == 'GET':
        return render_template('admin/import_is_elements.html')

    # POST: handle uploaded file
    file = request.files.get('file')
    if not file or file.filename == '':
        flash('Please choose a file to upload.', 'error')
        return redirect(url_for('admin.import_is_elements'))

    filename = file.filename.lower()
    success_count = 0
    error_count = 0
    error_details = []

    def create_element_from_row(row_dict, row_num):
        nonlocal success_count, error_count
        try:
            # For the new version we require CSV/Excel headers to match the database column names exactly.
            # Use only the exact column names from the SQL schema (lower-cased).
            
            def to_int(val):
                if val is None or str(val).strip() == '':
                    return None
                try:
                    return int(float(str(val)))
                except (ValueError, TypeError):
                    return None

            element = ISElement(
                name=(row_dict.get('name') or '').strip() or None,
                family=(row_dict.get('family') or '').strip() or None,
                group=(row_dict.get('group') or '').strip() or None,
                mge_type=(row_dict.get('mge_type') or '').strip() or None,
                host=(row_dict.get('host') or '').strip() or None,
                accession_number=(row_dict.get('accession_number') or '').strip() or None,
                iso=(row_dict.get('iso') or '').strip() or None,
                origin=(row_dict.get('origin') or '').strip() or None,
                tam=(row_dict.get('tam') or '').strip() or None,
                le_cleavage_site=(row_dict.get('le_cleavage_site') or '').strip() or None,
                re_cleavage_site=(row_dict.get('re_cleavage_site') or '').strip() or None,
                orf1=(row_dict.get('orf1') or '').strip() or None,
                orf2=(row_dict.get('orf2') or '').strip() or None,
                related_element_s=(row_dict.get('related_element_s') or '').strip() or None,
                isoform=(row_dict.get('isoform') or '').strip() or None,
                synomyns=(row_dict.get('synomyns') or '').strip() or None,
                transposition=(row_dict.get('transposition') or '').strip() or None,
                length=to_int(row_dict.get('length')),
                is_length=to_int(row_dict.get('is_length')),
                orf_number=to_int(row_dict.get('orf_number')),
                orf_1=to_int(row_dict.get('orf_1')),
                orf_1_length=to_int(row_dict.get('orf_1_length')),
                orf_1_begin=to_int(row_dict.get('orf_1_begin')),
                orf_1_end=to_int(row_dict.get('orf_1_end')),
                orf_1_strand=(row_dict.get('orf_1_strand') or '').strip() or None,
                fusion_orf_1=(row_dict.get('fusion_orf_1') or '').strip() or None,
                orf_1_function=(row_dict.get('orf_1_function') or '').strip() or None,
                orf_1_chemistry=(row_dict.get('orf_1_chemistry') or '').strip() or None,
                orf_1_sequence=(row_dict.get('orf_1_sequence') or '').strip() or None,
                orf_2=to_int(row_dict.get('orf_2')),
                orf_2_length=to_int(row_dict.get('orf_2_length')),
                orf_2_begin=to_int(row_dict.get('orf_2_begin')),
                orf_2_end=to_int(row_dict.get('orf_2_end')),
                orf_2_strand=(row_dict.get('orf_2_strand') or '').strip() or None,
                fusion_orf_2=(row_dict.get('fusion_orf_2') or '').strip() or None,
                orf_2_function=(row_dict.get('orf_2_function') or '').strip() or None,
                orf_2_chemistry=(row_dict.get('orf_2_chemistry') or '').strip() or None,
                orf_2_sequence=(row_dict.get('orf_2_sequence') or '').strip() or None,
                is_sequence=(row_dict.get('is_sequence') or '').strip() or None,
                comment=(row_dict.get('comment') or '').strip() or None,
                references=(row_dict.get('references') or '').strip() or None,
                status='approved',
                submitter_id=current_user.id,
                submitter_first_name=getattr(current_user, 'first_name', None),
                submitter_last_name=getattr(current_user, 'last_name', None),
                submitter_email=current_user.email,
                submitter_institution=getattr(current_user, 'institution', None),
                submitter_department=getattr(current_user, 'department', None),
                submitter_country=getattr(current_user, 'country', None),
                submitter_telephone=getattr(current_user, 'telephone', None),
                submitter_postal_address=getattr(current_user, 'postal_address', None),
                submitter_postal_code=getattr(current_user, 'postal_code', None),
                submission_date=datetime.now(timezone.utc)
            )

            db.session.add(element)
            success_count += 1
        except Exception as e:
            error_count += 1
            error_details.append(f"Row {row_num}: {str(e)}")

    try:
        if filename.endswith('.csv'):
            import csv, io
            text_stream = io.TextIOWrapper(file.stream, encoding='utf-8-sig')
            reader = csv.DictReader(text_stream)
            # normalize headers to lower-case
            reader.fieldnames = [h.lower() for h in reader.fieldnames] if reader.fieldnames else []
            # validate headers: require exact DB column names per new version
            # For admin imports we only require the minimal field set; submitter info is filled by the
            # current admin user. Other columns are optional and will be filled when present.
            required = {'name'}
            present = set(h for h in reader.fieldnames if h)
            missing = sorted(required - present)
            if missing:
                flash('Import failed: missing required columns: ' + ', '.join(missing), 'error')
                return redirect(url_for('admin.import_is_elements'))
            for i, row in enumerate(reader, 2): # Start from row 2 for error reporting
                # lower keys
                row_lower = {k.lower(): v for k, v in row.items() if k}
                create_element_from_row(row_lower, i)
        elif filename.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except Exception:
                flash('openpyxl is required to import .xlsx files. Please install it or upload a CSV file.', 'error')
                return redirect(url_for('admin.import_is_elements'))
            wb = load_workbook(filename=file, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter)
            if not headers:
                flash('Invalid Excel file: missing header row.', 'error')
                return redirect(url_for('admin.import_is_elements'))
            headers = [str(h).strip().lower() if h is not None else '' for h in headers]
            # validate headers for exact DB column names
            # For admin imports we only require the minimal field set; submitter info is filled by the
            # current admin user. Other columns are optional and will be filled when present.
            required = {'name'}
            present = set(h for h in headers if h)
            missing = sorted(required - present)
            if missing:
                flash('Import failed: missing required columns: ' + ', '.join(missing), 'error')
                return redirect(url_for('admin.import_is_elements'))
            for i, values in enumerate(rows_iter, 2): # Start from row 2 for error reporting
                row_dict = {headers[i]: (values[i] if i < len(values) else None) for i in range(len(headers))}
                # cast non-str values to str for uniform handling
                row_norm = {k: ('' if v is None else str(v)) for k, v in row_dict.items()}
                create_element_from_row(row_norm, i)
        else:
            flash('Unsupported file type. Please upload a CSV or XLSX file.', 'error')
            return redirect(url_for('admin.import_is_elements'))

        db.session.commit()

        # admin log
        AdminLog.log_action(
            user=current_user,
            action='batch_import_is_elements',
            resource_type='is_element',
            details={'success': success_count, 'error': error_count, 'error_details': error_details},
            request=request
        )

        if error_count > 0:
            flash(f'Import finished with errors. Success: {success_count}, Failed: {error_count}. See details below.', 'warning')
            # You can pass error_details to the template to display them
            return render_template('admin/import_is_elements.html', error_details=error_details)
        else:
            flash(f'Import finished successfully. Success: {success_count}.', 'success')
        
        return redirect(url_for('admin.is_elements'))
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Batch import failed: {str(e)}', exc_info=True)
        flash(f'Import failed: {str(e)}', 'error')
        return redirect(url_for('admin.import_is_elements'))

@admin.route('/is-element/add', methods=['GET', 'POST'])
@login_required
def add_is_element():
    """Add IS Element"""
    if not current_user.has_admin_permission():
        flash('您没有权限访问此页面。', 'error')
        return redirect(url_for('admin.login'))
    
    # Check that admin profile is complete
    if not current_user.first_name or not current_user.last_name or not current_user.email or not current_user.institution or not current_user.country:
        flash('请先完善个人资料（姓名、邮箱、机构、国家）后再提交数据。', 'warning')
        return redirect(url_for('auth.profile'))
    
    from app.forms.submission import ISElementSubmissionForm
    form = ISElementSubmissionForm()
    
    # Pre-fill admin info on GET
    if not form.is_submitted():
        form.first_name.data = current_user.first_name
        form.last_name.data = current_user.last_name
        form.email.data = current_user.email
        form.institution.data = current_user.institution
        form.department.data = current_user.department
        form.country.data = current_user.country or 'China'
        form.telephone.data = current_user.telephone
    
    if form.validate_on_submit():
        element = ISElement()
        
        # IS element fields
        element.name = form.name.data
        element.family = form.family.data
        element.group = form.group.data
        element.mge_type = form.mge_type.data
        element.host = form.host.data
        element.accession_number = form.accession_number.data
        element.origin = form.origin.data
        element.transposition = form.transposition.data
        element.is_length = form.is_length.data
        # left_flank/right_flank are read-only @property, inferred from is_sequence
        element.le_cleavage_site = form.le_cleavage_site.data
        element.re_cleavage_site = form.re_cleavage_site.data
        element.is_sequence = form.is_sequence.data
        element.orf_number = form.orf_number.data
        
        # ORF1 fields
        element.orf_1 = form.orf_1.data
        element.orf_1_length = form.orf_1_length.data
        element.orf_1_begin = form.orf_1_begin.data
        element.orf_1_end = form.orf_1_end.data
        element.orf_1_strand = form.orf_1_strand.data
        element.fusion_orf_1 = form.fusion_orf_1.data
        element.orf_1_function = form.orf_1_function.data
        element.orf_1_chemistry = form.orf_1_chemistry.data
        element.orf_1_sequence = form.orf_1_sequence.data
        
        # ORF2 fields
        element.orf_2 = form.orf_2.data
        element.orf_2_length = form.orf_2_length.data
        element.orf_2_begin = form.orf_2_begin.data
        element.orf_2_end = form.orf_2_end.data
        element.orf_2_strand = form.orf_2_strand.data
        element.fusion_orf_2 = form.fusion_orf_2.data
        element.orf_2_function = form.orf_2_function.data
        element.orf_2_chemistry = form.orf_2_chemistry.data
        element.orf_2_sequence = form.orf_2_sequence.data
        
        # Other fields
        element.orf1 = form.orf1.data
        element.orf2 = form.orf2.data
        element.tam = form.tam.data
        element.synomyns = form.synomyns.data
        element.iso = form.iso.data
        element.related_element_s = form.related_element_s.data
        element.isoform = form.isoform.data
        element.length = form.length.data
        element.comment = form.comment.data
        element.references = form.references.data
        
        # Force submitter fields from current admin profile
        element.submitter_first_name = current_user.first_name
        element.submitter_last_name = current_user.last_name
        element.submitter_institution = current_user.institution
        element.submitter_department = current_user.department
        element.submitter_postal_address = current_user.postal_address
        element.submitter_postal_code = current_user.postal_code
        element.submitter_country = current_user.country
        element.submitter_email = current_user.email
        element.submitter_telephone = current_user.telephone
        
        # Admin-added entries are auto-approved
        element.status = 'approved'
        element.submitter_id = current_user.id
        element.submission_date = datetime.now(timezone.utc)
        
        try:
            db.session.add(element)
            db.session.commit()
            
            # Audit log
            AdminLog.log_action(
                user=current_user,
                action='create_is_element',
                resource_type='is_element',
                resource_id=element.id,
                details={'name': element.name, 'family': element.family},
                request=request
            )
            
            flash(f'IS元素 "{element.name}" 添加成功！', 'success')
            return redirect(url_for('admin.is_elements'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败：{str(e)}', 'error')
    
    return render_template('admin/is_element_form.html', 
                         form=form, 
                         title='Add IS Element',
                         action_url=url_for('admin.add_is_element'))

@admin.route('/is-element/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_is_element(id):
    """编辑IS元素"""
    if not current_user.has_admin_permission():
        flash('您没有权限访问此页面。', 'error')
        return redirect(url_for('admin.login'))
    
    from app.forms.submission import AdminISElementForm
    element = ISElement.query.get_or_404(id)
    form = AdminISElementForm(obj=element)
    
    if form.validate_on_submit():
        # Save old data for audit log
        old_data = element.to_dict()
        
        # Update fields, skip read-only properties
        excluded_fields = ['left_flank', 'right_flank']  # read-only @property
        for field_name, field in form._fields.items():
            if field_name not in ['csrf_token', 'submit'] and field_name not in excluded_fields:
                if hasattr(element, field_name):
                    setattr(element, field_name, field.data)
        
        element.updated_at = datetime.now(timezone.utc)
        
        try:
            db.session.commit()
            
            # Audit log
            AdminLog.log_action(
                user=current_user,
                action='edit_is_element',
                resource_type='is_element',
                resource_id=element.id,
                details={
                    'name': element.name,
                    'changes': f'从 {old_data.get("name", "")} 修改为 {element.name}'
                },
                request=request
            )
            
            flash(f'IS元素 "{element.name}" 更新成功！', 'success')
            return redirect(url_for('admin.is_element_view', id=element.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
    
    # Delete form instance
    delete_form = DeleteForm()
    
    return render_template('admin/is_element_form.html', 
                         form=form, 
                         element=element,
                         title=f'编辑IS元素 - {element.name}',
                         action_url=url_for('admin.edit_is_element', id=id),
                         delete_form=delete_form)

@admin.route('/is-element/<int:id>/delete', methods=['POST'])
@login_required
def delete_is_element(id):
    """删除IS元素"""
    if not current_user.has_admin_permission():
        flash('您没有权限访问此页面。', 'error')
        return redirect(url_for('admin.login'))
    
    # Validate CSRF token
    form = DeleteForm()
    if not form.validate_on_submit():
        flash('删除失败：安全验证未通过。', 'error')
        return redirect(url_for('admin.is_elements'))
    
    element = ISElement.query.get_or_404(id)
    
    try:
        element_name = element.name
        
        # Record deletion log
        log = AdminLog(
            user_id=current_user.id,
            action='delete_is_element',
            resource_type='is_element',
            resource_id=element.id,
            details={'name': element_name},
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        db.session.add(log)
        
        # Delete the element
        db.session.delete(element)
        db.session.commit()
        
        flash(f'IS元素 "{element_name}" 已删除。', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'删除IS元素失败: {str(e)}', exc_info=True)
        
        # Specific error messages by error type
        if 'IntegrityError' in str(e):
            if 'cannot be null' in str(e):
                flash('删除失败：存在相关联的数据记录，请先处理相关数据。', 'error')
            elif 'FOREIGN KEY constraint fails' in str(e):
                flash('删除失败：该IS元素被其他数据引用，无法删除。', 'error')
            else:
                flash('删除失败：数据完整性约束冲突。', 'error')
        else:
            flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('admin.is_elements'))

@admin.route('/batch-action', methods=['POST'])
@login_required
def batch_action():
    """批量操作IS元素"""
    if not current_user.has_admin_permission():
        return jsonify({'success': False, 'message': '没有权限'}), 403
    
    action = request.form.get('action')
    comment = request.form.get('comment', '')
    
    # Get ID list, support various formats
    ids = request.form.getlist('ids')
    if not ids:
        # Try array-format IDs
        ids = []
        for key in request.form.keys():
            if key.startswith('ids['):
                ids.append(request.form.get(key))
    
    # Debug logging
    current_app.logger.info(f'批量操作请求: action={action}, ids={ids}, comment={comment}')
    current_app.logger.info(f'请求表单数据: {dict(request.form)}')
    current_app.logger.info(f'所有表单键: {list(request.form.keys())}')
    
    if not action or not ids:
        current_app.logger.warning(f'参数不完整: action={action}, ids={ids}')
        return jsonify({'success': False, 'message': f'参数不完整: action={action}, ids={ids}'}), 400
    
    try:
        # Validate IDs
        element_ids = [int(id) for id in ids]
        elements = ISElement.query.filter(ISElement.id.in_(element_ids)).all()
        
        if len(elements) != len(element_ids):
            return jsonify({'success': False, 'message': '部分IS元素不存在'}), 400
        
        success_count = 0
        
        for element in elements:
            try:
                if action in ['approve', 'pending', 'reject']:
                    # Status update operations
                    old_status = element.status
                    if action == 'approve':
                        element.status = 'approved'
                    elif action == 'pending':
                        element.status = 'pending'
                    elif action == 'reject':
                        element.status = 'rejected'
                    
                    # Audit log
                    log = AdminLog(
                        user_id=current_user.id,
                        action=f'batch_{action}_is_element',
                        resource_type='is_element',
                        resource_id=element.id,
                        details={
                            'name': element.name,
                            'old_status': old_status,
                            'new_status': element.status,
                            'comment': comment
                        },
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent')
                    )
                    db.session.add(log)
                    success_count += 1
                    
                elif action == 'delete':
                    # Deletion operation
                    element_name = element.name
                    log = AdminLog(
                        user_id=current_user.id,
                        action='batch_delete_is_element',
                        resource_type='is_element',
                        resource_id=element.id,
                        details={
                            'name': element_name,
                            'comment': comment
                        },
                        ip_address=request.remote_addr,
                        user_agent=request.headers.get('User-Agent')
                    )
                    db.session.add(log)
                    db.session.delete(element)
                    success_count += 1
                    
            except Exception as e:
                current_app.logger.warning(f'批量操作元素 {element.id} 失败: {str(e)}')
                continue
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'成功处理 {success_count} 个IS元素',
            'processed': success_count
        })
        
    except ValueError:
        return jsonify({'success': False, 'message': 'ID格式错误'}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'批量操作失败: {str(e)}', exc_info=True)
        return jsonify({'success': False, 'message': f'操作失败: {str(e)}'}), 500

@admin.route('/batch-export', methods=['POST'])
@login_required
def batch_export():
    """批量导出IS元素（支持CSV和FASTA格式）"""
    if not current_user.has_admin_permission():
        flash('您没有权限访问此页面。', 'error')
        return redirect(url_for('admin.login'))
    
    ids = request.form.getlist('ids')
    export_format = request.form.get('format', 'csv').lower()
    
    if not ids:
        flash('请选择要导出的IS元素。', 'error')
        return redirect(url_for('admin.is_elements'))
    
    try:
        from flask import Response
        from app.utils.fasta_generator import generate_fasta
        import csv
        import io
        from datetime import datetime, timezone
        
        # Validate IDs
        element_ids = [int(id) for id in ids]
        elements = ISElement.query.filter(ISElement.id.in_(element_ids)).all()
        
        if export_format == 'fasta':
            # 生成FASTA格式
            fasta_content = generate_fasta(elements, include_orfs=True)
            
            return Response(
                fasta_content,
                mimetype='text/plain',
                headers={
                    'Content-Disposition': f'attachment; filename=is_elements_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.fasta'
                }
            )
        else:
            # 生成CSV格式（完整数据）
            output = io.StringIO()
            writer = csv.writer(output)
            
            # 写入完整表头
            headers = [
                'Name', 'Family', 'Group', 'MGE Type', 'Related Element(s)', 'Isoform',
                'Synomyns', 'ISO', 'Accession Number', 'Transposition', 'Origin', 'Host',
                'IS Length', 'LE Cleavage Site', 'RE Cleavage Site', 'TAM',
                'IS Sequence', 'ORF Number', 'Length',
                'ORF_1', 'ORF_1_Length', 'ORF_1_Begin', 'ORF_1_End', 'ORF_1_Strand',
                'Fusion_ORF_1', 'ORF_1_Function', 'ORF_1_Chemistry', 'ORF_1_Sequence',
                'ORF_2', 'ORF_2_Length', 'ORF_2_Begin', 'ORF_2_End', 'ORF_2_Strand',
                'Fusion_ORF_2', 'ORF_2_Function', 'ORF_2_Chemistry', 'ORF_2_Sequence',
                'ORF1', 'ORF2', 'Comment', 'References', 'Status', 'Submission Date'
            ]
            writer.writerow(headers)
        
            
            # 写入数据行
            for element in elements:
                row = [
                    element.name or '',
                    element.family or '',
                    element.group or '',
                    element.mge_type or '',
                    element.related_element_s or '',
                    element.isoform or '',
                    element.synomyns or '',
                    element.iso or '',
                    element.accession_number or '',
                    element.transposition or '',
                    element.origin or '',
                    element.host or '',
                    element.is_length or '',
                    element.le_cleavage_site or '',
                    element.re_cleavage_site or '',
                    element.tam or '',
                    element.is_sequence or '',
                    element.orf_number or '',
                    element.length or '',
                    element.orf_1 or '',
                    element.orf_1_length or '',
                    element.orf_1_begin or '',
                    element.orf_1_end or '',
                    element.orf_1_strand or '',
                    element.fusion_orf_1 or '',
                    element.orf_1_function or '',
                    element.orf_1_chemistry or '',
                    element.orf_1_sequence or '',
                    element.orf_2 or '',
                    element.orf_2_length or '',
                    element.orf_2_begin or '',
                    element.orf_2_end or '',
                    element.orf_2_strand or '',
                    element.fusion_orf_2 or '',
                    element.orf_2_function or '',
                    element.orf_2_chemistry or '',
                    element.orf_2_sequence or '',
                    element.orf1 or '',
                    element.orf2 or '',
                    element.comment or '',
                    element.references or '',
                    element.status or '',
                    element.submission_date.strftime('%Y-%m-%d %H:%M:%S') if element.submission_date else ''
                ]
                writer.writerow(row)
            
            # 记录导出日志
            log = AdminLog(
                user_id=current_user.id,
                action=f'batch_export_is_elements_{export_format}',
                resource_type='is_element',
                details={
                    'count': len(elements),
                    'format': export_format,
                    'ids': element_ids[:10]  # 只记录前10个ID
                },
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent')
            )
            db.session.add(log)
            db.session.commit()
            
            # 创建响应
            output.seek(0)
            response = Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename=is_elements_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                }
            )
            
            return response
        
    except ValueError:
        flash('ID格式错误。', 'error')
        return redirect(url_for('admin.is_elements'))
    except Exception as e:
        current_app.logger.error(f'导出失败: {str(e)}', exc_info=True)
        flash(f'导出失败：{str(e)}', 'error')
        return redirect(url_for('admin.is_elements'))

@admin.route('/view-is-element/<int:id>')
@login_required
def view_is_element(id):
    """查看IS元素详情（AJAX接口）"""
    if not current_user.has_admin_permission():
        return {'error': '没有权限'}, 403
    
    element = ISElement.query.get_or_404(id)
    
    # Render detail HTML fragment
    return render_template('admin/is_element_detail.html', element=element)

@admin.route('/is-element/<int:id>/view')
@login_required
def is_element_view(id):
    """查看IS元素完整详情页面（管理员）"""
    if not current_user.has_admin_permission():
        flash('您没有权限访问此页面。', 'error')
        return redirect(url_for('admin.login'))
    
    element = ISElement.query.get_or_404(id)
    
    # Admin detail page template
    return render_template('admin/is_element_view.html', element=element)


@admin.route('/statistics')
@login_required
def statistics():
    """数据库统计页面（管理员）"""
    if not current_user.has_admin_permission():
        flash('您没有权限访问此页面。', 'error')
        return redirect(url_for('admin.login'))
    
    from sqlalchemy import func
    from datetime import datetime, timezone, timedelta
    
    # 基本统计
    stats = {
        'total_elements': ISElement.query.filter_by(status='approved').count(),
        'pending_elements': ISElement.query.filter_by(status='pending').count(),
        'total_articles': KnowledgeArticle.query.filter_by(status='published').count(),
        'total_users': User.query.filter_by(is_active=True).count()
    }
    
    # IS家族分布
    family_distribution = db.session.query(
        ISElement.family,
        func.count(ISElement.id).label('count')
    ).filter_by(status='approved')\
     .group_by(ISElement.family)\
     .order_by(func.count(ISElement.id).desc()).all()
    
    # MGE类型分布
    mge_type_distribution = db.session.query(
        ISElement.mge_type,
        func.count(ISElement.id).label('count')
    ).filter_by(status='approved')\
     .filter(ISElement.mge_type.isnot(None))\
     .group_by(ISElement.mge_type)\
     .order_by(func.count(ISElement.id).desc()).all()
    
    # 宿主分布（Top 10）
    host_distribution = db.session.query(
        ISElement.host,
        func.count(ISElement.id).label('count')
    ).filter_by(status='approved')\
     .filter(ISElement.host.isnot(None))\
     .group_by(ISElement.host)\
     .order_by(func.count(ISElement.id).desc())\
     .limit(10).all()
    
    # 长度分布统计
    length_ranges = [
        (0, 500), (500, 1000), (1000, 1500), 
        (1500, 2000), (2000, 3000), (3000, float('inf'))
    ]
    length_distribution = []
    
    for min_len, max_len in length_ranges:
        if max_len == float('inf'):
            count = ISElement.query.filter(
                ISElement.status == 'approved',
                ISElement.is_length >= min_len
            ).count()
            label = f'{min_len}bp+'
        else:
            count = ISElement.query.filter(
                ISElement.status == 'approved',
                ISElement.is_length >= min_len,
                ISElement.is_length < max_len
            ).count()
            label = f'{min_len}-{max_len}bp'
        
        length_distribution.append({'label': label, 'count': count})
    
    # 最近7天的搜索趋势
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    search_trends = db.session.query(
        func.date(SearchLog.created_at).label('date'),
        func.count(SearchLog.id).label('count')
    ).filter(SearchLog.created_at >= week_ago)\
     .group_by(func.date(SearchLog.created_at))\
     .order_by(func.date(SearchLog.created_at)).all()
    
    # 热门搜索词
    popular_searches = db.session.query(
        SearchLog.search_term,
        func.count(SearchLog.id).label('count')
    ).filter(SearchLog.created_at >= week_ago)\
     .group_by(SearchLog.search_term)\
     .order_by(func.count(SearchLog.id).desc())\
     .limit(10).all()
    
    return render_template('main/statistics.html',
                         stats=stats,
                         family_distribution=family_distribution,
                         mge_type_distribution=mge_type_distribution,
                         host_distribution=host_distribution,
                         length_distribution=length_distribution,
                         search_trends=search_trends,
                         popular_searches=popular_searches)

# ============== Admin Management Routes (Root Only) ==============

@admin.route('/admins')
@login_required
def manage_admins():
    """管理员账号管理（仅root可访问）"""
    if not current_user.is_root():
        flash('只有ROOT用户可以访问此页面。', 'error')
        return redirect(url_for('admin.index'))
    
    # 获取所有管理员和root用户
    admins = User.query.filter(User.role.in_(['root', 'admin'])).order_by(
        User.role.desc(),  # root排在前面
        User.created_at.desc()
    ).all()
    
    return render_template('admin/admins.html', admins=admins)

@admin.route('/admins/create', methods=['GET', 'POST'])
@login_required
def create_admin_user():
    """创建管理员账号（仅root可访问）"""
    if not current_user.is_root():
        flash('只有ROOT用户可以创建管理员账号。', 'error')
        return redirect(url_for('admin.index'))
    
    from app.forms.admin_management import CreateAdminForm
    form = CreateAdminForm()
    
    if form.validate_on_submit():
        admin_user = User(
            username=form.username.data,
            email=form.email.data if form.email.data else None,
            role='admin',
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            institution=form.institution.data,
            department=form.department.data,
            country=form.country.data,
            is_active=True
        )
        admin_user.password = form.password.data
        
        try:
            db.session.add(admin_user)
            db.session.commit()
            
            # Audit log
            AdminLog.log_action(
                user=current_user,
                action='create_admin',
                resource_type='user',
                resource_id=admin_user.id,
                details={
                    'username': admin_user.username, 
                    'email': admin_user.email or 'Not set',
                    'has_email': bool(admin_user.email)
                },
                request=request
            )
            
            flash(f'管理员账号 "{admin_user.username}" 创建成功！', 'success')
            return redirect(url_for('admin.manage_admins'))
        except Exception as e:
            db.session.rollback()
            flash(f'创建失败：{str(e)}', 'error')
    
    return render_template('admin/create_admin.html', form=form)

@admin.route('/admins/<int:admin_id>/disable', methods=['POST'])
@login_required
def disable_admin(admin_id):
    """禁用管理员账号（仅root可访问）"""
    if not current_user.is_root():
        flash('只有ROOT用户可以禁用管理员账号。', 'error')
        return redirect(url_for('admin.index'))
    
    admin_user = User.query.get_or_404(admin_id)
    
    if admin_user.role == 'root':
        flash('不能禁用ROOT账号。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    if admin_user.role != 'admin':
        flash('只能禁用admin角色的用户。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    admin_user.is_active = False
    
    try:
        db.session.commit()
        
        # Audit log
        AdminLog.log_action(
            user=current_user,
            action='disable_admin',
            resource_type='user',
            resource_id=admin_user.id,
            details={'username': admin_user.username},
            request=request
        )
        
        flash(f'管理员 "{admin_user.username}" 已被禁用。', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'禁用失败：{str(e)}', 'error')
    
    return redirect(url_for('admin.manage_admins'))

@admin.route('/admins/<int:admin_id>/enable', methods=['POST'])
@login_required
def enable_admin(admin_id):
    """启用管理员账号（仅root可访问）"""
    if not current_user.is_root():
        flash('只有ROOT用户可以启用管理员账号。', 'error')
        return redirect(url_for('admin.index'))
    
    admin_user = User.query.get_or_404(admin_id)
    
    if admin_user.role != 'admin':
        flash('只能启用admin角色的用户。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    admin_user.is_active = True
    
    try:
        db.session.commit()
        
        # Audit log
        AdminLog.log_action(
            user=current_user,
            action='enable_admin',
            resource_type='user',
            resource_id=admin_user.id,
            details={'username': admin_user.username},
            request=request
        )
        
        flash(f'管理员 "{admin_user.username}" 已被启用。', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'启用失败：{str(e)}', 'error')
    
    return redirect(url_for('admin.manage_admins'))

@admin.route('/admins/<int:admin_id>/delete', methods=['POST'])
@login_required
def delete_admin(admin_id):
    """删除管理员账号（仅root可访问）"""
    if not current_user.is_root():
        flash('只有ROOT用户可以删除管理员账号。', 'error')
        return redirect(url_for('admin.index'))
    
    admin_user = User.query.get_or_404(admin_id)
    
    if admin_user.role == 'root':
        flash('不能删除ROOT账号。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    if admin_user.role != 'admin':
        flash('只能删除admin角色的用户。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    # 检查是否有关联数据
    if admin_user.submitted_elements.count() > 0:
        flash(f'管理员 "{admin_user.username}" 有关联的提交数据，不能删除。请先转移或删除相关数据。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    if admin_user.reviewed_elements.count() > 0:
        flash(f'管理员 "{admin_user.username}" 有关联的审核数据，不能删除。请先转移或删除相关数据。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    if admin_user.articles.count() > 0:
        flash(f'管理员 "{admin_user.username}" 有关联的文章，不能删除。请先转移或删除相关数据。', 'error')
        return redirect(url_for('admin.manage_admins'))
    
    username = admin_user.username
    
    try:
        # Audit log（在删除前记录）
        AdminLog.log_action(
            user=current_user,
            action='delete_admin',
            resource_type='user',
            resource_id=admin_user.id,
            details={'username': username, 'email': admin_user.email},
            request=request
        )
        
        db.session.delete(admin_user)
        db.session.commit()
        
        flash(f'管理员 "{username}" 已被删除。', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败：{str(e)}', 'error')
    
    return redirect(url_for('admin.manage_admins'))


@admin.route('/download-requests')
@login_required
def download_requests():
    """查看下载申请列表"""
    if not current_user.has_admin_permission():
        flash('需要管理员权限。', 'error')
        return redirect(url_for('main.index'))
    
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', 'pending')
    
    query = DownloadRequest.query
    
    # 如果是普通admin，只看分配给自己的
    if current_user.role == 'admin':
        query = query.filter_by(assigned_admin_id=current_user.id)
    
    # 状态过滤
    if status_filter and status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    requests = query.order_by(DownloadRequest.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('admin/download_requests.html',
                         requests=requests,
                         status_filter=status_filter)


@admin.route('/download-requests/<int:request_id>')
@login_required
def download_request_detail(request_id):
    """查看下载申请详情"""
    if not current_user.has_admin_permission():
        flash('需要管理员权限。', 'error')
        return redirect(url_for('main.index'))
    
    download_req = DownloadRequest.query.get_or_404(request_id)
    
    # 普通admin只能看分配给自己的
    if current_user.role == 'admin' and download_req.assigned_admin_id != current_user.id:
        flash('无权查看此申请。', 'error')
        return redirect(url_for('admin.download_requests'))
    
    # 获取申请的IS元素列表
    items = DownloadRequestItem.query.filter_by(request_id=request_id).all()
    elements = [db.session.get(ISElement, item.is_element_id) for item in items]

    return render_template('admin/download_request_detail.html',
                         download_request=download_req,
                         elements=elements)


@admin.route('/download-requests/<int:request_id>/approve', methods=['POST'])
@login_required
def approve_download_request(request_id):
    """审核通过下载申请"""
    if not current_user.has_admin_permission():
        flash('需要管理员权限。', 'error')
        return redirect(url_for('main.index'))
    
    download_req = DownloadRequest.query.get_or_404(request_id)
    
    # 普通admin只能审核分配给自己的
    if current_user.role == 'admin' and download_req.assigned_admin_id != current_user.id:
        flash('无权审核此申请。', 'error')
        return redirect(url_for('admin.download_requests'))
    
    comment = request.form.get('comment', '').strip()
    
    try:
        download_req.status = 'approved'
        download_req.reviewed_at = datetime.now(timezone.utc)
        download_req.reviewer_comment = comment
        
        db.session.commit()
        
        # 记录日志
        AdminLog.log_action(
            user=current_user,
            action='approve_download_request',
            resource_type='download_request',
            resource_id=request_id,
            details={'comment': comment},
            request=request
        )
        
        # 发送邮件通知（包含数据文件）
        send_download_approval_email(download_req)
        
        flash('申请已批准，邮件已发送。', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'操作失败：{str(e)}', 'error')
    
    return redirect(url_for('admin.download_request_detail', request_id=request_id))


@admin.route('/download-requests/<int:request_id>/reject', methods=['POST'])
@login_required
def reject_download_request(request_id):
    """拒绝下载申请"""
    if not current_user.has_admin_permission():
        flash('需要管理员权限。', 'error')
        return redirect(url_for('main.index'))
    
    download_req = DownloadRequest.query.get_or_404(request_id)
    
    # 普通admin只能审核分配给自己的
    if current_user.role == 'admin' and download_req.assigned_admin_id != current_user.id:
        flash('无权审核此申请。', 'error')
        return redirect(url_for('admin.download_requests'))
    
    comment = request.form.get('comment', '').strip()
    if not comment:
        flash('拒绝时必须填写理由。', 'error')
        return redirect(url_for('admin.download_request_detail', request_id=request_id))
    
    try:
        download_req.status = 'rejected'
        download_req.reviewed_at = datetime.now(timezone.utc)
        download_req.reviewer_comment = comment
        
        db.session.commit()
        
        # 记录日志
        AdminLog.log_action(
            user=current_user,
            action='reject_download_request',
            resource_type='download_request',
            resource_id=request_id,
            details={'comment': comment},
            request=request
        )
        
        flash('申请已拒绝。', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'操作失败：{str(e)}', 'error')
    
    return redirect(url_for('admin.download_request_detail', request_id=request_id))


@admin.route('/download-requests/delete', methods=['POST'])
@login_required
def delete_download_requests():
    """批量删除下载申请"""
    if not current_user.has_admin_permission():
        return jsonify({'success': False, 'message': '需要管理员权限'}), 403
    
    try:
        data = request.get_json()
        ids = data.get('ids', [])
        
        if not ids:
            return jsonify({'success': False, 'message': '未选择要删除的记录'}), 400
        
        # 删除记录
        deleted_count = 0
        for req_id in ids:
            download_req = db.session.get(DownloadRequest, req_id)
            if download_req:
                # 记录日志
                AdminLog.log_action(
                    user=current_user,
                    action='delete_download_request',
                    resource_type='download_request',
                    resource_id=req_id,
                    details={
                        'applicant': f"{download_req.first_name} {download_req.last_name}",
                        'status': download_req.status,
                        'items_count': download_req.items.count()
                    },
                    request=request
                )
                
                db.session.delete(download_req)
                deleted_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'成功删除 {deleted_count} 条记录'
        })
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'删除下载申请失败: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500


def send_download_approval_email(download_req):
    """发送批准邮件，根据申请格式附带数据文件（CSV或FASTA）"""
    from flask_mail import Mail, Message
    from app.utils.fasta_generator import generate_fasta
    import csv
    import io
    
    mail = Mail(current_app)
    
    # 获取申请的IS元素
    items = DownloadRequestItem.query.filter_by(request_id=download_req.id).all()
    elements = [db.session.get(ISElement, item.is_element_id) for item in items]

    # 根据申请的格式生成文件
    export_format = download_req.format or 'csv'
    
    if export_format == 'fasta':
        # 生成FASTA格式
        file_data = generate_fasta(elements, include_orfs=True)
        file_extension = 'fasta'
        content_type = 'text/plain'
    else:
        # 生成CSV格式（包含核心字段和ORF详细信息）
        csv_buffer = io.StringIO()
        csv_writer = csv.writer(csv_buffer)
        
        # 写入表头（核心字段 + ORF详细信息）
        csv_writer.writerow([
            'Name', 'Family', 'Group', 'MGE Type', 'Related Element(s)', 'Isoform',
            'Synomyns', 'ISO', 'Accession Number', 'Transposition', 'Origin', 'Host',
            'IS Length', 'LE Cleavage Site', 'RE Cleavage Site', 'TAM',
            'IS Sequence', 'ORF Number', 'Length',
            'ORF_1', 'ORF_1_Length', 'ORF_1_Begin', 'ORF_1_End', 'ORF_1_Strand',
            'Fusion_ORF_1', 'ORF_1_Function', 'ORF_1_Chemistry', 'ORF_1_Sequence',
            'ORF_2', 'ORF_2_Length', 'ORF_2_Begin', 'ORF_2_End', 'ORF_2_Strand',
            'Fusion_ORF_2', 'ORF_2_Function', 'ORF_2_Chemistry', 'ORF_2_Sequence',
            'ORF1', 'ORF2', 'Comment', 'References'
        ])
        
        # 写入数据
        for elem in elements:
            csv_writer.writerow([
                elem.name or '',
                elem.family or '',
                elem.group or '',
                elem.mge_type or '',
                elem.related_element_s or '',
                elem.isoform or '',
                elem.synomyns or '',
                elem.iso or '',
                elem.accession_number or '',
                elem.transposition or '',
                elem.origin or '',
                elem.host or '',
                elem.is_length or '',
                elem.le_cleavage_site or '',
                elem.re_cleavage_site or '',
                elem.tam or '',
                elem.is_sequence or '',
                elem.orf_number or '',
                elem.length or '',
                elem.orf_1 or '',
                elem.orf_1_length or '',
            elem.orf_1_begin or '',
            elem.orf_1_end or '',
            elem.orf_1_strand or '',
            elem.fusion_orf_1 or '',
            elem.orf_1_function or '',
                elem.orf_1_chemistry or '',
                elem.orf_1_sequence or '',
                elem.orf_2 or '',
                elem.orf_2_length or '',
                elem.orf_2_begin or '',
                elem.orf_2_end or '',
                elem.orf_2_strand or '',
                elem.fusion_orf_2 or '',
                elem.orf_2_function or '',
                elem.orf_2_chemistry or '',
                elem.orf_2_sequence or '',
                elem.orf1 or '',
                elem.orf2 or '',
                elem.comment or '',
                elem.references or ''
            ])
        
        file_data = csv_buffer.getvalue()
        file_extension = 'csv'
        content_type = 'text/csv'
    
    # 创建邮件
    msg = Message(
        subject='Download Request Approved - proISDB',
        recipients=[download_req.email],
        body=f'''Dear {download_req.first_name} {download_req.last_name},

Your download request has been approved.

Request Details:
- Request ID: {download_req.id}
- Number of elements: {len(elements)}
- Export format: {export_format.upper()}
- Approved by: {current_user.username}
- Approval time: {download_req.reviewed_at.strftime("%Y-%m-%d %H:%M:%S") if download_req.reviewed_at else "N/A"}

The requested data is attached as a {export_format.upper()} file.

Best regards,
proISDB Team
'''
    )
    
    # 附加文件
    msg.attach(
        filename=f'proisdb_data_{download_req.id}.{file_extension}',
        content_type=content_type,
        data=file_data
    )
    
    # 发送邮件
    try:
        mail.send(msg)
    except Exception as e:
        current_app.logger.error(f'Failed to send email: {str(e)}')
        raise


# ============== BLAST数据库管理 ==============

@admin.route('/blast-database', methods=['GET'])
@login_required
def blast_database():
    """BLAST数据库管理页面"""
    if not current_user.is_admin():
        flash('权限不足', 'danger')
        return redirect(url_for('main.index'))
    
    import os
    import subprocess
    from datetime import datetime, timezone
    
    # 获取数据库信息
    blast_db_dir = os.path.join(current_app.root_path, '..', 'blast_db')
    nucl_db_path = os.path.join(blast_db_dir, 'is_elements')
    prot_db_path = os.path.join(blast_db_dir, 'is_elements_prot')
    
    db_info = {
        'nucleotide': {
            'exists': False,
            'update_time': None,
            'sequences': 0,
            'file_size': 0
        },
        'protein': {
            'exists': False,
            'update_time': None,
            'sequences': 0,
            'file_size': 0
        }
    }
    
    # 检查核酸数据库
    nucl_files = [f"{nucl_db_path}.{ext}" for ext in ['nhr', 'nin', 'nsq']]
    if all(os.path.exists(f) for f in nucl_files):
        db_info['nucleotide']['exists'] = True
        # 获取更新时间
        mtime = os.path.getmtime(nucl_files[0])
        db_info['nucleotide']['update_time'] = datetime.fromtimestamp(mtime)
        # 获取文件大小
        total_size = sum(os.path.getsize(f) for f in nucl_files if os.path.exists(f))
        db_info['nucleotide']['file_size'] = total_size / 1024 / 1024  # MB
        
        # 获取序列数量
        try:
            result = subprocess.run(
                ['blastdbcmd', '-db', nucl_db_path, '-info'],
                capture_output=True,
                text=True,
                timeout=5
            )
            if result.returncode == 0:
                for line in result.stdout.split('\n'):
                    if 'sequences' in line.lower():
                        import re
                        match = re.search(r'(\d+)\s+sequences', line, re.IGNORECASE)
                        if match:
                            db_info['nucleotide']['sequences'] = int(match.group(1))
        except:
            pass
    
    # 检查蛋白数据库
    prot_files = [f"{prot_db_path}.{ext}" for ext in ['phr', 'pin', 'psq']]
    if all(os.path.exists(f) for f in prot_files):
        db_info['protein']['exists'] = True
        mtime = os.path.getmtime(prot_files[0])
        db_info['protein']['update_time'] = datetime.fromtimestamp(mtime)
        total_size = sum(os.path.getsize(f) for f in prot_files if os.path.exists(f))
        db_info['protein']['file_size'] = total_size / 1024 / 1024  # MB
        
        try:
            result = subprocess.run(
                ['blastdbcmd', '-db', prot_db_path, '-info'],
                capture_output=True,
                text=True,
                timeout=5
            )
            if result.returncode == 0:
                for line in result.stdout.split('\n'):
                    if 'sequences' in line.lower():
                        import re
                        match = re.search(r'(\d+)\s+sequences', line, re.IGNORECASE)
                        if match:
                            db_info['protein']['sequences'] = int(match.group(1))
        except:
            pass
    
    # 获取IS元素总数
    total_is_elements = ISElement.query.filter(
        ISElement.is_sequence.isnot(None),
        ISElement.is_sequence != ''
    ).count()
    
    # 检查BLAST+是否安装
    blast_installed = False
    try:
        result = subprocess.run(['blastn', '-version'], capture_output=True, timeout=5)
        blast_installed = result.returncode == 0
    except:
        pass
    
    return render_template(
        'admin/blast_database.html',
        db_info=db_info,
        total_is_elements=total_is_elements,
        blast_installed=blast_installed
    )


@admin.route('/blast-database/rebuild', methods=['POST'])
@login_required
def rebuild_blast_database():
    """重建BLAST数据库"""
    if not current_user.is_admin():
        return jsonify({'success': False, 'message': '权限不足'}), 403
    
    import os
    import subprocess
    
    try:
        # 运行构建脚本
        script_path = os.path.join(current_app.root_path, '..', 'build_blast_db.py')
        
        result = subprocess.run(
            ['python', script_path],
            capture_output=True,
            text=True,
            timeout=300  # 5分钟超时
        )
        
        if result.returncode == 0:
            # 记录管理员操作
            AdminLog.log_action(
                user=current_user,
                action='rebuild_blast_database',
                resource_type='system',
                details={'output': result.stdout[:500]},
                request=request
            )
            
            return jsonify({
                'success': True,
                'message': 'BLAST数据库重建成功',
                'output': result.stdout
            })
        else:
            return jsonify({
                'success': False,
                'message': 'BLAST数据库重建失败',
                'error': result.stderr
            }), 500
            
    except subprocess.TimeoutExpired:
        return jsonify({
            'success': False,
            'message': '重建超时（>5分钟），请检查数据量或服务器性能'
        }), 500
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'重建失败: {str(e)}'
        }), 500
